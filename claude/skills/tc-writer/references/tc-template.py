"""
TC xlsx 생성 뼈대 코드
─────────────────────
사용법:
  1. tc_data 리스트에 TC 내용 채우기
  2. OUTPUT_PATH 지정
  3. python tc-template.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 출력 경로 ──────────────────────────────────────────────
OUTPUT_PATH = "/mnt/user-data/outputs/TC_문서.xlsx"

# ── 색상 팔레트 ────────────────────────────────────────────
HEADER_BG  = "1F2D3D"   # 헤더 배경 (진한 네이비)
SUBHDR_BG  = "2E4057"   # 그룹 구분행 배경 (중간 네이비)
ALT_BG     = "F4F6F9"   # 홀수 TC 행 배경 (연한 회색)
WHITE      = "FFFFFF"
BORDER_CLR = "BEC8D2"
BRAND      = "FE6F47"   # P0 강조 (브랜드 오렌지)

# ── TC 데이터 ──────────────────────────────────────────────
# type: "group" → 모듈 구분행 (label만 필요)
# type: "tc"    → 실제 TC 행
tc_data = [
    {"type": "group", "label": "01. 모듈명"},
    {
        "type": "tc",
        "id": "tc-001",
        "title": "TC 제목",
        "priority": "P0",       # P0 / P1 / P2
        "mode": "공통",          # 공통 / 관리자 / 사용자 등
        "precond": "전제조건 없음 (초기 상태에서 수행 가능)",
        "data": "없음",
        "steps": "1. 진입경로 > 메뉴 클릭\n2. 버튼 클릭\n3. 결과 확인",
        "expected": "• 형태: 토스트 메시지 (화면 우상단)\n• 문구: \"메시지 텍스트\"\n• 3초 후 자동 사라짐",
    },
    # 필요한 만큼 추가...
]

# ── 컬럼 정의 (순서·너비 고정) ─────────────────────────────
COLUMNS = [
    ("TC-ID",        12),
    ("제목",          28),
    ("우선순위",       9),
    ("모드",           9),
    ("전제조건",       28),
    ("테스트 데이터",  24),
    ("수행 절차",      38),
    ("기대 결과",      38),
    ("실제 결과",      22),
    ("Pass / Fail",   12),
]

# ── 스타일 헬퍼 ───────────────────────────────────────────
_thin = Side(style="thin", color=BORDER_CLR)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(size=9, bold=False, color="1A1A1A", italic=False):
    return Font(name="Arial", size=size, bold=bold, color=color, italic=italic)

def _align(h="left"):
    return Alignment(wrap_text=True, vertical="top", horizontal=h)

# ── 빌드 ──────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "TC"

# 헤더 행
for i, (name, width) in enumerate(COLUMNS, start=1):
    c = ws.cell(row=1, column=i, value=name)
    c.font = _font(size=10, bold=True, color="FFFFFF")
    c.fill = _fill(HEADER_BG)
    c.border = BORDER
    c.alignment = _align("center")
    ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[1].height = 28

# TC 행
row = 2
for item in tc_data:
    if item["type"] == "group":
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
        c = ws.cell(row=row, column=1, value=item["label"])
        c.font = _font(size=10, bold=True, color="FFFFFF")
        c.fill = _fill(SUBHDR_BG)
        c.border = BORDER
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.row_dimensions[row].height = 20
        row += 1
        continue

    bg = WHITE if row % 2 == 0 else ALT_BG
    values = [
        item["id"], item["title"], item["priority"], item["mode"],
        item["precond"], item["data"], item["steps"], item["expected"],
        "(수행 후 기입)", "",
    ]

    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _fill(bg)
        c.border = BORDER

        if col == 3 and val == "P0":            # P0 강조
            c.font = _font(bold=True, color=BRAND)
            c.alignment = _align("center")
        elif col == 9:                           # 실제 결과 힌트
            c.font = _font(color="AAAAAA", italic=True)
            c.alignment = _align("center")
        elif col in (1, 3, 4, 10):              # ID / 우선순위 / 모드 / Pass
            c.font = _font()
            c.alignment = _align("center")
        else:
            c.font = _font()
            c.alignment = _align("left")

    ws.row_dimensions[row].height = 72
    row += 1

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

wb.save(OUTPUT_PATH)
print("saved:", OUTPUT_PATH)